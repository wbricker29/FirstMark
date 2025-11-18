#!/usr/bin/env python3
"""
Export Helpers for Airtable

Export Airtable data to multiple formats:
- CSV export with proper encoding
- JSON export with nested structures
- Markdown table generation
- Excel export with formatting
- Handles large datasets efficiently

Usage:
    from scripts.export_helpers import ExportHelpers

    exporter = ExportHelpers(records, table_schema)
    exporter.to_excel("output.xlsx", include_formatting=True)
"""

import json
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime


class ExportHelpers:
    """Export Airtable data to various formats"""

    def __init__(
        self,
        records: List[Dict[str, Any]],
        table_schema: Optional[Dict[str, Any]] = None,
    ):
        """
        Initialize exporter with records

        Args:
            records: List of Airtable records
            table_schema: Optional table schema for enhanced exports
        """
        self.records = records
        self.schema = table_schema
        self.field_names = self._extract_field_names()

    def to_csv(self, filename: str, encoding: str = "utf-8") -> str:
        """
        Export records to CSV

        Args:
            filename: Output file path
            encoding: Character encoding (default: utf-8)

        Returns:
            Path to created file
        """
        print(f"📊 Exporting {len(self.records)} records to CSV...")

        with open(filename, "w", newline="", encoding=encoding) as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=["id"] + list(self.field_names))
            writer.writeheader()

            for record in self.records:
                row = {"id": record.get("id", "")}
                fields = record.get("fields", {})

                for field_name in self.field_names:
                    value = fields.get(field_name, "")

                    # Handle complex types
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)

                    row[field_name] = value

                writer.writerow(row)

        print(f"✓ CSV export complete: {filename}")
        return filename

    def to_json(
        self, filename: str, pretty: bool = True, include_metadata: bool = True
    ) -> str:
        """
        Export records to JSON

        Args:
            filename: Output file path
            pretty: Pretty print with indentation
            include_metadata: Include export metadata

        Returns:
            Path to created file
        """
        print(f"📊 Exporting {len(self.records)} records to JSON...")

        export_data = {"records": self.records}

        if include_metadata:
            export_data["metadata"] = {
                "exported_at": datetime.now().isoformat(),
                "record_count": len(self.records),
                "schema": self.schema,
            }

        with open(filename, "w", encoding="utf-8") as jsonfile:
            if pretty:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
            else:
                json.dump(export_data, jsonfile, ensure_ascii=False)

        print(f"✓ JSON export complete: {filename}")
        return filename

    def to_markdown(
        self, filename: str, max_col_width: int = 50, include_id: bool = False
    ) -> str:
        """
        Export records to Markdown table

        Args:
            filename: Output file path
            max_col_width: Maximum column width for text truncation
            include_id: Include record ID column

        Returns:
            Path to created file
        """
        print(f"📊 Exporting {len(self.records)} records to Markdown...")

        with open(filename, "w", encoding="utf-8") as mdfile:
            # Header
            mdfile.write("# Airtable Export\n\n")
            mdfile.write(
                f"**Exported:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            )
            mdfile.write(f"**Records:** {len(self.records)}\n\n")

            # Table header
            columns = (
                ["ID"] + list(self.field_names)
                if include_id
                else list(self.field_names)
            )
            mdfile.write("| " + " | ".join(columns) + " |\n")
            mdfile.write("| " + " | ".join(["---"] * len(columns)) + " |\n")

            # Table rows
            for record in self.records:
                row_values = []

                if include_id:
                    row_values.append(record.get("id", ""))

                fields = record.get("fields", {})
                for field_name in self.field_names:
                    value = fields.get(field_name, "")

                    # Format value
                    if isinstance(value, list):
                        if all(isinstance(x, str) for x in value):
                            value = ", ".join(value)
                        else:
                            value = str(len(value)) + " items"
                    elif isinstance(value, dict):
                        value = "Object"
                    else:
                        value = str(value)

                    # Truncate long values
                    if len(value) > max_col_width:
                        value = value[: max_col_width - 3] + "..."

                    # Escape pipe characters
                    value = value.replace("|", "\\|")

                    row_values.append(value)

                mdfile.write("| " + " | ".join(row_values) + " |\n")

        print(f"✓ Markdown export complete: {filename}")
        return filename

    def to_excel(
        self, filename: str, include_formatting: bool = True, freeze_header: bool = True
    ) -> str:
        """
        Export records to Excel with formatting

        Args:
            filename: Output file path (.xlsx)
            include_formatting: Apply cell formatting
            freeze_header: Freeze header row

        Returns:
            Path to created file

        Note: Requires openpyxl package
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("❌ openpyxl not installed. Install with: pip install openpyxl")
            return ""

        print(f"📊 Exporting {len(self.records)} records to Excel...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Airtable Export"

        # Header row
        headers = ["ID"] + list(self.field_names)
        ws.append(headers)

        # Format header
        if include_formatting:
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        # Data rows
        for record in self.records:
            row = [record.get("id", "")]
            fields = record.get("fields", {})

            for field_name in self.field_names:
                value = fields.get(field_name, "")

                # Handle complex types
                if isinstance(value, list):
                    if all(isinstance(x, str) for x in value):
                        value = ", ".join(value)
                    else:
                        value = f"{len(value)} items"
                elif isinstance(value, dict):
                    value = str(value)

                row.append(value)

            ws.append(row)

        # Auto-adjust column widths
        if include_formatting:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = "A2"

        wb.save(filename)
        print(f"✓ Excel export complete: {filename}")
        return filename

    def to_html(
        self,
        filename: str,
        include_styles: bool = True,
        table_class: str = "airtable-export",
    ) -> str:
        """
        Export records to HTML table

        Args:
            filename: Output file path
            include_styles: Include basic CSS styles
            table_class: CSS class for table

        Returns:
            Path to created file
        """
        print(f"📊 Exporting {len(self.records)} records to HTML...")

        with open(filename, "w", encoding="utf-8") as htmlfile:
            # HTML header
            htmlfile.write("<!DOCTYPE html>\n")
            htmlfile.write("<html>\n<head>\n")
            htmlfile.write("<meta charset='UTF-8'>\n")
            htmlfile.write("<title>Airtable Export</title>\n")

            # CSS styles
            if include_styles:
                htmlfile.write("<style>\n")
                htmlfile.write(f".{table_class} {{\n")
                htmlfile.write("  border-collapse: collapse;\n")
                htmlfile.write("  width: 100%;\n")
                htmlfile.write("  font-family: Arial, sans-serif;\n")
                htmlfile.write("}\n")
                htmlfile.write(f".{table_class} th, .{table_class} td {{\n")
                htmlfile.write("  border: 1px solid #ddd;\n")
                htmlfile.write("  padding: 8px;\n")
                htmlfile.write("  text-align: left;\n")
                htmlfile.write("}\n")
                htmlfile.write(f".{table_class} th {{\n")
                htmlfile.write("  background-color: #366092;\n")
                htmlfile.write("  color: white;\n")
                htmlfile.write("  font-weight: bold;\n")
                htmlfile.write("}\n")
                htmlfile.write(f".{table_class} tr:nth-child(even) {{\n")
                htmlfile.write("  background-color: #f2f2f2;\n")
                htmlfile.write("}\n")
                htmlfile.write("</style>\n")

            htmlfile.write("</head>\n<body>\n")
            htmlfile.write("<h1>Airtable Export</h1>\n")
            htmlfile.write(
                f"<p>Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>\n"
            )
            htmlfile.write(f"<p>Records: {len(self.records)}</p>\n")

            # Table
            htmlfile.write(f"<table class='{table_class}'>\n")

            # Header
            htmlfile.write("<thead>\n<tr>\n")
            htmlfile.write("<th>ID</th>\n")
            for field_name in self.field_names:
                htmlfile.write(f"<th>{self._escape_html(field_name)}</th>\n")
            htmlfile.write("</tr>\n</thead>\n")

            # Body
            htmlfile.write("<tbody>\n")
            for record in self.records:
                htmlfile.write("<tr>\n")
                htmlfile.write(f"<td>{self._escape_html(record.get('id', ''))}</td>\n")

                fields = record.get("fields", {})
                for field_name in self.field_names:
                    value = fields.get(field_name, "")

                    if isinstance(value, list):
                        value = ", ".join(str(x) for x in value)
                    elif isinstance(value, dict):
                        value = "Object"

                    htmlfile.write(f"<td>{self._escape_html(str(value))}</td>\n")

                htmlfile.write("</tr>\n")
            htmlfile.write("</tbody>\n")

            htmlfile.write("</table>\n")
            htmlfile.write("</body>\n</html>")

        print(f"✓ HTML export complete: {filename}")
        return filename

    def _extract_field_names(self) -> List[str]:
        """Extract unique field names from records"""
        field_names_set = set()

        for record in self.records:
            fields = record.get("fields", {})
            field_names_set.update(fields.keys())

        return sorted(list(field_names_set))

    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters"""
        return (
            str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;")
        )


# Example usage
if __name__ == "__main__":
    # Example records
    sample_records = [
        {
            "id": "rec1",
            "fields": {
                "Name": "John Doe",
                "Email": "john@example.com",
                "Status": "Active",
                "Tags": ["customer", "vip"],
            },
        },
        {
            "id": "rec2",
            "fields": {
                "Name": "Jane Smith",
                "Email": "jane@example.com",
                "Status": "Pending",
                "Tags": ["prospect"],
            },
        },
    ]

    exporter = ExportHelpers(sample_records)

    # Export to different formats
    exporter.to_csv("export.csv")
    exporter.to_json("export.json")
    exporter.to_markdown("export.md")
    exporter.to_html("export.html")

    # Note: Excel export requires openpyxl
    # exporter.to_excel('export.xlsx')

    print("\n✓ All exports complete!")
